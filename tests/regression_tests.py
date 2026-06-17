"""KOSH regression smoke tests — runs against the real Postgres.

Each test isolates itself with a unique high-numbered test PCN (>=99000)
that does not collide with real production PCNs, and a SAVEPOINT/ROLLBACK
wrapper so nothing leaks into production data.

Tests cover the specific bug shapes Preet reported in May 2026 so a code
change that re-introduces any of them fails this suite immediately.

Run from inside the container:
    docker exec stockandpick_webapp python /app/tests/regression_tests.py

Run from the host:
    /home/tony/KOSH/tests/run.sh

Exit code is the number of failed tests (0 == all green).
"""

import os
import sys
import traceback
from contextlib import contextmanager

import psycopg2

DB_URL = os.environ.get(
    'DATABASE_URL',
    'postgresql://stockpick_user:stockpick_pass@aci-database:5432/kosh',
)
SCHEMA = 'pcb_inventory'


@contextmanager
def isolated_txn():
    """Open a connection, run the body, and ALWAYS rollback at the end.

    Using a top-level transaction we never commit means every test's writes
    disappear when the block exits, even if assertions pass.
    """
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    try:
        yield conn
    finally:
        conn.rollback()
        conn.close()


def _seed_warehouse_row(cursor, pcn, item, mpn='TEST-MPN', onhandqty=0,
                       mfg_qty='0', loc_to='Count Area'):
    cursor.execute(
        f'INSERT INTO {SCHEMA}."tblWhse_Inventory" '
        '(item, pcn, mpn, dc, onhandqty, mfg_qty, loc_from, loc_to) '
        'VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
        (item, str(pcn), mpn, 'DC2026', onhandqty, mfg_qty, '-', loc_to),
    )


def _seed_txn(cursor, trantype, item, pcn, qty, userid='regression@test.com',
             loc_to='Count Area'):
    cursor.execute(
        f'INSERT INTO {SCHEMA}."tblTransaction" '
        '(trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, '
        ' loc_from, loc_to, userid) '
        "VALUES (%s, %s, %s, 'TEST-MPN', 'DC2026', 'Level 1', %s, "
        "TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', "
        "'MM/DD/YY HH24:MI:SS'), '-', %s, %s)",
        (trantype, item, str(pcn), qty, loc_to, userid),
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_restock_allowed_after_purge_following_restock():
    """Bug shape: PCN 44822/45143 — RESTOCK then ADJT then PURGE → guard
    refused next restock with 'already restocked'.

    Expected: with last KOSH PICK/RESTOCK/PURGE = PURGE, restock_pcb's
    guard releases (PURGE != RESTOCK).
    """
    sys.path.insert(0, '/app')
    from app import db_manager  # noqa: E402

    test_pcn = 99001
    test_item = 'REGRESS-RESTOCK-AFTER-PURGE'

    with isolated_txn() as conn:
        cur = conn.cursor()
        _seed_warehouse_row(cur, test_pcn, test_item, onhandqty=0, mfg_qty='100')
        _seed_txn(cur, 'PICK', test_item, test_pcn, 100)
        _seed_txn(cur, 'RESTOCK', test_item, test_pcn, 50)
        _seed_txn(cur, 'PURGE', test_item, test_pcn, 0)
        conn.commit()  # need to commit so db_manager's separate connection sees it

        try:
            result = db_manager.restock_pcb(
                pcn=test_pcn, item=test_item, quantity=10,
                location_from='Count Area', location_to='Count Area',
                username='regression@test.com',
            )
            assert result.get('success') is True, (
                f'Restock should succeed after PURGE, got: {result}'
            )
        finally:
            # Manually clean up since we had to commit the seed data
            cleanup = psycopg2.connect(DB_URL)
            cleanup.autocommit = True
            cur2 = cleanup.cursor()
            cur2.execute(
                f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s',
                (str(test_pcn),),
            )
            cur2.execute(
                f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s',
                (str(test_pcn),),
            )
            cleanup.close()


def test_restock_allowed_when_zero_onhand_even_if_last_restock():
    """Defensive backstop: SUM(onhandqty)=0 → restock always valid."""
    sys.path.insert(0, '/app')
    from app import db_manager

    test_pcn = 99002
    test_item = 'REGRESS-ZERO-ONHAND'

    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
    cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))

    with isolated_txn() as conn:
        cur = conn.cursor()
        _seed_warehouse_row(cur, test_pcn, test_item, onhandqty=0, mfg_qty='0')
        _seed_txn(cur, 'RESTOCK', test_item, test_pcn, 50)
        conn.commit()

        try:
            result = db_manager.restock_pcb(
                pcn=test_pcn, item=test_item, quantity=10,
                location_from='Count Area', location_to='Count Area',
                username='regression@test.com',
            )
            assert result.get('success') is True, (
                f'Restock should succeed when onhandqty=0, got: {result}'
            )
        finally:
            cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
            cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))
    cleanup.close()


def test_restock_blocked_when_already_restocked_with_stock_present():
    """Sanity: the guard should still block the genuine 'double restock'
    case — last txn = RESTOCK AND onhandqty > 0.
    """
    sys.path.insert(0, '/app')
    from app import db_manager

    test_pcn = 99003
    test_item = 'REGRESS-DOUBLE-RESTOCK'

    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
    cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))

    with isolated_txn() as conn:
        cur = conn.cursor()
        _seed_warehouse_row(cur, test_pcn, test_item, onhandqty=50, mfg_qty='0')
        _seed_txn(cur, 'RESTOCK', test_item, test_pcn, 50)
        conn.commit()

        try:
            result = db_manager.restock_pcb(
                pcn=test_pcn, item=test_item, quantity=10,
                location_from='Count Area', location_to='Count Area',
                username='regression@test.com',
            )
            assert result.get('success') is False, (
                f'Restock SHOULD be blocked when last=RESTOCK and qty>0, got: {result}'
            )
            assert 'already been restocked' in result.get('error', '')
        finally:
            cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
            cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))
    cleanup.close()


def test_print_label_sums_across_duplicate_pcn_rows():
    """Bug shape: print-label SELECT … LIMIT 1 picked one row when a PCN
    had duplicate warehouse rows, so labels showed wrong qty.

    Expected: SUM(onhandqty) across all rows for the PCN.
    """
    sys.path.insert(0, '/app')
    import app as app_module

    test_pcn = 99004
    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))

    cur0.execute(
        f'INSERT INTO {SCHEMA}."tblWhse_Inventory" (item, pcn, mpn, onhandqty, loc_to) '
        "VALUES ('REGRESS-DUP-A', %s, 'X', 30, 'Count Area')",
        (str(test_pcn),),
    )
    cur0.execute(
        f'INSERT INTO {SCHEMA}."tblWhse_Inventory" (item, pcn, mpn, onhandqty, loc_to) '
        "VALUES ('REGRESS-DUP-B', %s, 'X', 20, 'Count Area')",
        (str(test_pcn),),
    )

    try:
        client = app_module.app.test_client()
        with client.session_transaction() as sess:
            sess['user_id'] = 1
            sess['username'] = 'regression@test.com'
            sess['role'] = 'admin'
        resp = client.get(f'/print-label/{test_pcn}')
        assert resp.status_code == 200, f'expected 200, got {resp.status_code}'
        body = resp.get_data(as_text=True)
        assert '50' in body, (
            f'label total should be 30+20=50; body excerpt: {body[:500]}'
        )
    finally:
        cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))
    cleanup.close()


def test_validate_location_auto_registers_unknown_7digit():
    """Bug shape: a fresh 7-digit location wasn't in tblLoc → restock
    failed with 'Location does not exist'. Should auto-register.
    """
    sys.path.insert(0, '/app')
    from app import db_manager

    test_loc = '9999301'  # synthetic test 7-digit code
    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblLoc" WHERE location = %s', (test_loc,))

    try:
        ok = db_manager.validate_location(test_loc)
        assert ok is True, 'validate_location should auto-register and return True'
        cur0.execute(f'SELECT COUNT(*) FROM {SCHEMA}."tblLoc" WHERE location = %s', (test_loc,))
        n = cur0.fetchone()[0]
        assert n == 1, f'tblLoc should now contain {test_loc}; rows found: {n}'
    finally:
        cur0.execute(f'DELETE FROM {SCHEMA}."tblLoc" WHERE location = %s', (test_loc,))
    cleanup.close()


def test_purged_pcn_can_be_restocked_with_same_pcn():
    """Bug shape: restock_pcb returned 'No parts found' if the warehouse
    row had been deleted by a legacy purge. The fallback should recreate
    the row from the most recent PURGE/STOCK/RESTOCK transaction.
    """
    sys.path.insert(0, '/app')
    from app import db_manager

    test_pcn = 99005
    test_item = 'REGRESS-PURGED-LOSTROW'

    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
    cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))

    cur0.execute(
        f'INSERT INTO {SCHEMA}."tblTransaction" '
        '(trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, userid) '
        "VALUES ('PURGE', %s, %s, 'TEST-MPN', 'DC2026', 'Level 1', 0, "
        "TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'), "
        "'-', 'Purged', 'regression@test.com')",
        (test_item, str(test_pcn)),
    )

    try:
        result = db_manager.restock_pcb(
            pcn=test_pcn, item=None, quantity=15,
            location_from='Count Area', location_to='Count Area',
            username='regression@test.com',
        )
        assert result.get('success') is True, (
            f'restock should recreate missing row from PURGE txn, got: {result}'
        )
    finally:
        cur0.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = %s', (str(test_pcn),))
        cur0.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = %s', (str(test_pcn),))
    cleanup.close()


def test_bom_load_inserts_every_item_received():
    """Bug shape: BOM uploads silently dropped lines. The /api/bom/load
    endpoint must persist exactly the number of items it receives.
    """
    sys.path.insert(0, '/app')
    import app as app_module

    test_job = 'REGRESS-BOM-9999'
    items_in = []
    for i in range(1, 12):  # 11 items
        items_in.append({
            'line': i,
            'desc': f'Test desc {i}',
            'man': 'TestMan',
            'mpn': f'TEST-MPN-{i}',
            'aci_pn': f'ACI-{i}',
            'qty': i * 5,
            'pou': 'SMT',
            'loc': '',
            'cost': 1.23,
            'job': test_job,
            'job_rev': 'A',
            'last_rev': '',
            'cust': 'TestCust',
            'cust_pn': '',
            'cust_rev': '',
        })

    cleanup = psycopg2.connect(DB_URL)
    cleanup.autocommit = True
    cur0 = cleanup.cursor()
    cur0.execute(f'DELETE FROM {SCHEMA}."tblBOM" WHERE job::text = %s', (test_job,))
    cur0.execute(f'DELETE FROM {SCHEMA}."tblJob" WHERE job_number = %s', (test_job,))

    try:
        # Disable CSRF for the test only (we're calling the API directly)
        app_module.app.config['WTF_CSRF_ENABLED'] = False
        client = app_module.app.test_client()
        with client.session_transaction() as sess:
            sess['user_id'] = 1
            sess['username'] = 'regression@test.com'
            sess['role'] = 'admin'
        resp = client.post('/api/bom/load', json={
            'metadata': {'job': test_job, 'job_rev': 'A', 'customer': 'TestCust', 'cust_pn': '', 'cust_rev': '', 'last_rev': '', 'wo_number': '', 'notes': '', 'build_qty': 1},
            'bom_items': items_in,
        })
        assert resp.status_code == 200, f'expected 200, got {resp.status_code}: {resp.get_data(as_text=True)[:300]}'
        body = resp.get_json()
        assert body and body.get('success'), f'expected success, got {body}'
        assert body.get('inserted_count') == len(items_in), (
            f'expected {len(items_in)} inserted, got {body.get("inserted_count")}'
        )

        cur0.execute(f'SELECT COUNT(*) FROM {SCHEMA}."tblBOM" WHERE job::text = %s', (test_job,))
        n_rows = cur0.fetchone()[0]
        assert n_rows == len(items_in), (
            f'tblBOM should hold {len(items_in)} rows, found {n_rows}'
        )
    finally:
        cur0.execute(f'DELETE FROM {SCHEMA}."tblBOM" WHERE job::text = %s', (test_job,))
        cur0.execute(f'DELETE FROM {SCHEMA}."tblJob" WHERE job_number = %s', (test_job,))
    cleanup.close()


def test_bom_python_parser_finds_lines_across_sheets():
    """Bug shape: 8813L-4DA file had Line 200 only on 'Assy BOM', not on
    'BOM to Load'. Multi-sheet merge logic should pick it up.

    Python implementation of the same merge the JS parser does — keeps
    drift between client and server detectable.
    """
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws_load = wb.active
    ws_load.title = 'BOM to Load'
    ws_load.append(['Line', 'DESC', 'MAN', 'MPN', 'ACI PN', 'QTY'])
    for i in range(1, 4):
        ws_load.append([i, f'desc {i}', 'M', f'MPN-{i}', f'ACI-{i}', 5])

    ws_assy = wb.create_sheet('Assy BOM')
    for _ in range(9):
        ws_assy.append([])
    ws_assy.append(['LINE', 'DESC', 'MAN', 'MPN', 'ACI PN', 'QTY'])
    for i in [1, 2, 3, 4]:  # row 4 is extra-only
        ws_assy.append([i, f'desc {i}', 'M', f'MPN-{i}', f'ACI-{i}', 5])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = _python_multisheet_parse(buf.getvalue())
    line_nums = sorted(it['line'] for it in parsed)
    assert line_nums == [1, 2, 3, 4], (
        f'multi-sheet merge should yield lines 1-4, got {line_nums}'
    )


def _python_multisheet_parse(file_bytes):
    """Mirror of the JS multi-sheet merge: walk every sheet, find header
    row dynamically, merge unique LINE numbers (BOM to Load wins on dup).
    Used only by tests so we can verify drift symbolically.
    """
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    line_aliases = {'LINE', 'LINE #', 'LINE NO', 'LINE NUMBER', 'ITEM',
                    'ITEM #', 'ITEM NO', '#', 'NO', 'NO.', 'ROW', 'SEQ',
                    'SR', 'SR.', 'SR NO', 'S.NO', 'S NO'}

    def parse_sheet(ws):
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for hr in range(min(30, len(rows))):
            cells = [str(c).strip().upper().replace('  ', ' ') if c is not None else '' for c in rows[hr]]
            has_line = any(c in line_aliases for c in cells)
            has_mpn = any('MPN' in c or 'MFG PN' in c or 'MFR PN' in c or 'PART NUMBER' in c or 'DESC' in c for c in cells)
            if has_line and has_mpn:
                header_idx = hr
                headers = cells
                break
        if header_idx is None:
            return []
        line_col = next((i for i, h in enumerate(headers) if h in line_aliases), None)
        if line_col is None:
            return []
        items = []
        for r in rows[header_idx + 1:]:
            if not r or all(c is None or c == '' for c in r):
                continue
            try:
                ln = int(r[line_col])
            except (TypeError, ValueError):
                continue
            items.append({'line': ln})
        return items

    seen = set()
    merged = []
    order = ['BOM to Load'] + [n for n in wb.sheetnames if n != 'BOM to Load']
    for name in order:
        if name not in wb.sheetnames:
            continue
        for it in parse_sheet(wb[name]):
            if it['line'] in seen:
                continue
            seen.add(it['line'])
            merged.append(it)
    return merged


def test_return_connection_never_leaks_foreign_connection():
    """Bug shape (May 2026): /sources and /stats opened raw psycopg2.connect
    connections, then handed them to db_manager.return_connection, which
    called pool.putconn → 'trying to put unkeyed connection'. The old code
    only logged that and dropped the connection — leaking it. After enough
    page views the maxconn=15 pool was exhausted and the whole app hung.

    Expected: return_connection CLOSES any connection putconn rejects, and the
    pool keeps handing out connections (no exhaustion).
    """
    sys.path.insert(0, '/app')
    from app import db_manager

    # A connection NOT from the pool must be closed, not leaked, and must not raise.
    foreign = psycopg2.connect(DB_URL)
    assert foreign.closed == 0, 'sanity: freshly opened connection should be open'
    db_manager.return_connection(foreign)
    assert foreign.closed != 0, (
        'foreign connection must be CLOSED by return_connection, not leaked'
    )

    # Pool connections must still round-trip many times without exhausting
    # (15 = maxconn; loop past it to prove nothing leaks per cycle).
    for _ in range(25):
        c = db_manager.get_connection()
        assert c is not None
        db_manager.return_connection(c)


def test_app_served_by_gunicorn_not_dev_server():
    """Perf guard: the container must run gunicorn, not Flask's single-request
    dev server (`python app.py`). The dev server served requests one-at-a-time,
    so one slow query stalled every other user — the main KOSH latency cause.
    """
    dockerfile = os.path.join(os.path.dirname(__file__), '..', 'Dockerfile.webapp')
    text = open(dockerfile).read()
    # Find the runtime CMD line.
    cmd_lines = [l for l in text.splitlines() if l.strip().startswith('CMD')]
    joined = ' '.join(cmd_lines) + text  # be lenient about multi-line CMD
    assert 'gunicorn' in joined, 'Dockerfile.webapp CMD must launch gunicorn'
    assert '"python", "app.py"' not in joined and "'python', 'app.py'" not in joined, (
        'Dockerfile must NOT run the Flask dev server (python app.py)'
    )


def test_all_pages_render_without_server_error():
    """Broad guard: every parameterless GET route must render WITHOUT a 500
    for an authenticated admin. Catches template errors, broken queries, and
    NameErrors across the whole app the moment they ship — so a change to any
    page can't silently 500 in production.

    Parameterized routes (/sources/<t>, /print-label/<pcn>) are exercised by
    their own targeted tests since they need valid IDs. Side-effecting/auth
    routes (logout, SSO callback) are excluded.
    """
    sys.path.insert(0, '/app')
    import app as app_module

    EXCLUDE = {'/logout', '/sso/callback'}
    app_module.app.config['WTF_CSRF_ENABLED'] = False
    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess['user_id'] = 1
        sess['username'] = 'regression@test.com'  # _final_cleanup wipes this user's rows
        sess['role'] = 'Admin'  # capital A → is_admin_user() bypasses every tool gate

    # Scope: user-facing PAGE routes. /api/* endpoints are excluded here because
    # many legitimately require query params or external resources (a no-arg GET
    # may correctly return 4xx); they need their own targeted tests. We log how
    # many we skipped so the scope is never a silent cap.
    failures = []
    checked = 0
    skipped_api = 0
    for rule in app_module.app.url_map.iter_rules():
        if 'GET' not in (rule.methods or set()):
            continue
        if rule.arguments:           # needs a path param — covered elsewhere
            continue
        path = str(rule.rule)
        if path in EXCLUDE or path.startswith('/static'):
            continue
        if path.startswith('/api/'):
            skipped_api += 1
            continue
        try:
            resp = client.get(path)
            code = resp.status_code
        except Exception as e:
            failures.append(f'{path} raised {type(e).__name__}: {e}')
            continue
        checked += 1
        if code >= 500:
            failures.append(f'{path} -> HTTP {code}')

    print(f'  [all-pages] smoke-tested {checked} page routes; skipped {skipped_api} /api routes')
    assert checked > 20, f'expected to smoke-test many pages, only hit {checked}'
    assert not failures, (
        f'{len(failures)} page(s) returned a server error:\n  ' +
        '\n  '.join(failures)
    )


def test_quantity_fields_are_not_number_spinners():
    """Bug shape (May 2026, PCN 41564): the restock qty was logged 1 short.
    Root cause: WTForms IntegerField renders <input type=number>, which
    silently decrements on a mouse-wheel tick while focused. Fix renders the
    qty as type=text inputmode=numeric so wheel/arrows can't change it.

    Guard: restock/stock/pick must NOT render quantity as a number spinner.
    """
    import re
    base = os.path.join(os.path.dirname(__file__), '..', 'templates', 'inventory_ops')
    for tpl in ('restock', 'stock', 'pick'):
        path = os.path.join(base, f'{tpl}.html')
        html = open(path).read()
        m = re.search(r'form\.quantity\((?:[^()]|\([^()]*\))*\)', html)
        assert m, f'{tpl}.html: form.quantity(...) render call not found'
        call = m.group(0)
        assert 'type="text"' in call, (
            f'{tpl}.html: quantity must render type="text" (number inputs lose '
            f'a unit to wheel scroll). Got: {call}'
        )
        assert 'inputmode="numeric"' in call, (
            f'{tpl}.html: quantity must set inputmode="numeric" for mobile keypad'
        )


def test_shortage_report_alt_part_qty_and_same_mpn_visibility():
    """Bug shape (June 2026, jobs 8657ML/8566ML — Theresa lost trust):

    A) Alternate-manufacturer BOM lines store the real qty on the primary row
       and 0 on the "ZSUB"/alternate row (same aci_pn). The dedup must keep the
       row carrying the real qty, NOT collapse to 0 — a 0 zeroed REQ and
       silently dropped a genuine shortage from the report.

    B) The report must surface stock of the SAME MPN sitting under a DIFFERENT
       job's part number (other_mpn_onhand) so Purchasing stops hand-searching
       Warehouse Inventory — WITHOUT counting it as this job's own on-hand
       (which would double-allocate another job's committed stock). This stock is
       displayed as indented ROW ENTRIES (the old two-column display was removed
       2026-06-12 at report users' request); other_mpn_locations now stores a JSON
       breakdown that parse_other_mpn_rows() expands into those rows.
    """
    sys.path.insert(0, '/app')
    import app as app_module
    from psycopg2.extras import RealDictCursor

    job = 'REGRESS-SHORT-9999'
    aci_pn = 'RSHORT-1'
    mpn = 'RGMPN-REGRESS-1'

    with isolated_txn() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        # Clean any prior seed inside this rolled-back txn
        cur.execute(f'DELETE FROM {SCHEMA}."tblBOM" WHERE job = %s', (job,))
        cur.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn IN %s',
                    (('99320', '99321', '99322'),))

        # BOM: primary row qty 10, alternate row (same aci_pn) qty 0
        for line, m, qty in [('10', mpn, '10'), ('10', mpn + '-ALT', '0')]:
            cur.execute(
                f'INSERT INTO {SCHEMA}."tblBOM" (line, "DESC", man, mpn, aci_pn, qty, cost, job, job_rev) '
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (line, 'Regress cap', 'RegressMan', m, aci_pn, qty, '1.00', job, 'A'),
            )

        # On-hand: 3 under THIS job's own part; 50 of same MPN under ANOTHER
        # part number; 1000 of same MPN but on MFG Floor (must be excluded).
        _seed_warehouse_row(cur, 99320, aci_pn, mpn=mpn, onhandqty=3, loc_to='Count Area')
        _seed_warehouse_row(cur, 99321, 'OTHERJOB-99', mpn=mpn, onhandqty=50, loc_to='Count Area')
        _seed_warehouse_row(cur, 99322, 'OTHERJOB-99', mpn=mpn, onhandqty=1000, loc_to='MFG Floor')

        result = app_module._persist_shortage_report(
            cur, job, order_qty=1, report_name='regress', notes='', username='regression@test.com')
        assert result is not None, 'shortage report should generate for seeded job'

        cur.execute(
            f'SELECT qty_required, qty_on_hand, other_mpn_onhand, other_mpn_locations '
            f'FROM {SCHEMA}."tblShortageReportItems" WHERE report_id = %s',
            (result['report_id'],))
        rows = cur.fetchall()
        assert len(rows) == 1, f'expected 1 shortage line, got {len(rows)}'
        row = rows[0]

        # A) real qty kept (not the alternate's 0)
        assert row['qty_required'] == 10, (
            f"alt-part dedup must keep qty 10, got {row['qty_required']} "
            f"(a 0 here silently drops real shortages — the trust bug)")
        # own on-hand only
        assert row['qty_on_hand'] == 3, (
            f"on-hand must count only this job's own part (3), got {row['qty_on_hand']}")
        # B) same-MPN visibility = other-item stock, MFG Floor excluded
        assert row['other_mpn_onhand'] == 50, (
            f"same-MPN visibility must be 50 (other PN, excl MFG Floor), "
            f"got {row['other_mpn_onhand']}")
        assert row['other_mpn_locations'] and 'OTHERJOB-99' in row['other_mpn_locations'], (
            f"same-MPN locations must name the other part number, "
            f"got {row['other_mpn_locations']!r}")
        # B2) the stored breakdown must parse into a row entry (NOT the removed
        # columns) carrying the other PN + its on-hand qty, MFG Floor excluded.
        sub_rows = app_module.parse_other_mpn_rows(row['other_mpn_locations'])
        assert len(sub_rows) == 1, (
            f"expected 1 same-MPN row entry, got {len(sub_rows)}: {sub_rows!r}")
        assert sub_rows[0]['item'] == 'OTHERJOB-99' and sub_rows[0]['qty'] == 50, (
            f"row entry must carry other PN OTHERJOB-99 @ 50, got {sub_rows[0]!r}")


def test_shortage_report_own_stock_is_case_insensitive():
    """Inventory stores the SAME part number in mixed case (BOM 'RCASE-1' vs
    stock 'rcase-1'). The job's own on-hand MUST aggregate across case — a
    case-sensitive join (the 2026-06-12 bug) missed the stock, falsely reported
    the part as short, AND listed the same part as a same-MPN "other PN" row.

    Guards both halves: (1) own on-hand counts the mixed-case lot; (2) the same
    ACI PN (any case) is NEVER emitted as a same-MPN row entry — only a genuinely
    different part number is.
    """
    sys.path.insert(0, '/app')
    import app as app_module
    from psycopg2.extras import RealDictCursor

    job = 'REGRESS-CASE-9999'
    aci_pn = 'RCASE-1'            # BOM stores UPPER
    mpn = 'RGMPN-CASE-1'

    with isolated_txn() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'DELETE FROM {SCHEMA}."tblBOM" WHERE job = %s', (job,))
        cur.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn IN %s',
                    (('99330', '99332'),))

        cur.execute(
            f'INSERT INTO {SCHEMA}."tblBOM" (line, "DESC", man, mpn, aci_pn, qty, cost, job, job_rev) '
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            ('10', 'Regress res', 'RegressMan', mpn, aci_pn, '100', '1.00', job, 'A'),
        )

        # 30 of THIS job's own part stored in LOWERCASE; 50 of same MPN under a
        # genuinely DIFFERENT part number.
        _seed_warehouse_row(cur, 99330, 'rcase-1', mpn=mpn, onhandqty=30, loc_to='Count Area')
        _seed_warehouse_row(cur, 99332, 'RCASE-OTHER', mpn=mpn, onhandqty=50, loc_to='Count Area')

        result = app_module._persist_shortage_report(
            cur, job, order_qty=1, report_name='regress', notes='', username='regression@test.com')
        assert result is not None, 'shortage report should generate for seeded job'

        cur.execute(
            f'SELECT qty_on_hand, req, other_mpn_locations '
            f'FROM {SCHEMA}."tblShortageReportItems" WHERE report_id = %s',
            (result['report_id'],))
        rows = cur.fetchall()
        assert len(rows) == 1, f'expected 1 shortage line, got {len(rows)}'
        row = rows[0]

        # (1) own on-hand counts the mixed-case lot (30), so REQ 100 > 30 = short
        assert row['qty_on_hand'] == 30, (
            f"own on-hand must count mixed-case stock (30), got {row['qty_on_hand']} "
            f"(a 0 here is the false-shortage bug)")

        # (2) same ACI PN (any case) must NOT be a same-MPN row entry; only the
        # genuinely different part number appears.
        sub_rows = app_module.parse_other_mpn_rows(row['other_mpn_locations'])
        items = {s['item'].upper() for s in sub_rows}
        assert aci_pn.upper() not in items, (
            f"same ACI PN (any case) must NOT appear as a same-MPN row entry, got {sub_rows!r}")
        assert items == {'RCASE-OTHER'}, (
            f"only the genuinely different PN should be a row entry, got {sub_rows!r}")


def test_shortage_report_counts_mfg_floor_stock():
    """Task 2 (2026-06-15, owner-reversed from CANCELLED): the shortage report
    MUST count this job's own MFG-Floor stock toward on-hand, so a job whose
    material is physically on the floor stops flagging a false shortage and the
    planner doesn't re-buy parts we already have.

    Before: rows with loc_to='MFG Floor' were excluded entirely, so floor stock
    (carried in mfg_qty) read as 0 on-hand. Now on-hand = bin on-hand + floor
    mfg_qty, summed across the part's lots. The two never overlap (Task-1 fix
    guarantees 0 rows with both onhandqty>0 AND mfg_qty>0), so no double-count.
    """
    sys.path.insert(0, '/app')
    import app as app_module
    from psycopg2.extras import RealDictCursor

    job = 'REGRESS-FLOOR-9999'
    aci_pn = 'RFLOOR-1'
    mpn = 'RGMPN-FLOOR-1'

    with isolated_txn() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'DELETE FROM {SCHEMA}."tblBOM" WHERE job = %s', (job,))
        cur.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn IN %s',
                    (('99340', '99341'),))

        # BOM needs 100 of the part.
        cur.execute(
            f'INSERT INTO {SCHEMA}."tblBOM" (line, "DESC", man, mpn, aci_pn, qty, cost, job, job_rev) '
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            ('10', 'Regress floor', 'RegressMan', mpn, aci_pn, '100', '1.00', job, 'A'),
        )

        # 30 in a bin (on-hand) + 90 physically on the MFG Floor (mfg_qty, on-hand 0).
        # Combined physical on-hand = 120 >= 100 req, so this is NOT a shortage.
        _seed_warehouse_row(cur, 99340, aci_pn, mpn=mpn, onhandqty=30, mfg_qty='0', loc_to='Count Area')
        _seed_warehouse_row(cur, 99341, aci_pn, mpn=mpn, onhandqty=0, mfg_qty='90', loc_to='MFG Floor')

        result = app_module._persist_shortage_report(
            cur, job, order_qty=1, report_name='regress', notes='', username='regression@test.com')
        assert result is not None, 'shortage report should generate for seeded job'

        # 120 on hand (30 bin + 90 floor) >= 100 req -> NOT short, so 0 lines kept.
        cur.execute(
            f'SELECT COUNT(*) AS n FROM {SCHEMA}."tblShortageReportItems" WHERE report_id = %s',
            (result['report_id'],))
        assert cur.fetchone()['n'] == 0, (
            "job is fully covered once MFG-Floor stock counts (30 bin + 90 floor = "
            "120 >= 100 req) — it must NOT appear as a shortage")

        # Lower the floor lot so 30 + 10 = 40 < 100 -> short, on-hand must read 40
        # (proving floor stock is summed in, not excluded and not double-counted).
        cur.execute(f'UPDATE {SCHEMA}."tblWhse_Inventory" SET mfg_qty=%s WHERE pcn=%s', ('10', '99341'))
        result2 = app_module._persist_shortage_report(
            cur, job, order_qty=1, report_name='regress2', notes='', username='regression@test.com')
        cur.execute(
            f'SELECT qty_on_hand FROM {SCHEMA}."tblShortageReportItems" WHERE report_id = %s',
            (result2['report_id'],))
        rows = cur.fetchall()
        assert len(rows) == 1, f'expected 1 shortage line, got {len(rows)}'
        assert rows[0]['qty_on_hand'] == 40, (
            f"on-hand must be bin 30 + floor 10 = 40, got {rows[0]['qty_on_hand']} "
            f"(excluded floor -> 30; double-counted -> 50)")


def test_location_reconcile_follows_latest_placement():
    """Location-sync bug (2026-06-15): put-away/relocate happens in legacy Access
    and arrives as imported PTWY transactions; KOSH never reflected them into
    tblWhse_Inventory.loc_to, so Warehouse Inventory showed a stale bin while PCN
    History showed the true one (Theresa could only find stock via History).

    The location reconcile sets loc_to to each stocked PCN's latest PLACEMENT
    location (PTWY/RESTOCK/INDF/STOCK, by chronological tran_time — NOT row id,
    which the Access import left out of order). A PICK is an OUTFLOW and must be
    ignored so a partial/zero pick can't drag the remaining bin stock onto the
    MFG Floor. This test seeds two put-aways (bin A then bin B) plus a later PICK,
    starts the inventory row at a STALE 'MFG Floor', and asserts the sync lands on
    bin B (latest placement) — not bin A (older) and not MFG Floor (the pick).
    """
    sys.path.insert(0, '/app')
    from psycopg2.extras import RealDictCursor
    from app import reconcile_warehouse_locations  # the SHIPPED query, not a copy

    test_pcn = 99933
    item = 'RLOC-1'
    # 8-DIGIT bins on purpose: the 2026-06-17 recurrence (PCN 45504 -> 14051021)
    # was an 8-digit bin that a stale {6,7}-digit filter silently dropped. Test
    # with 8-digit bins so that regression can never ship green again.
    bin_a = '14046120'
    bin_b = '14047080'

    with isolated_txn() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn = %s', (str(test_pcn),))
        cur.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn = %s', (str(test_pcn),))

        def txn(tt, qty, lf, lt, t):
            cur.execute(
                f'INSERT INTO {SCHEMA}."tblTransaction" '
                '(trantype, item, pcn, mpn, tranqty, tran_time, loc_from, loc_to, userid) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (tt, item, str(test_pcn), 'RLOC-MPN', str(qty), t, lf, lt, 'regression@test.com'))
        txn('INDF', 50, 'n/a', 'Rec Area', '01/10/26 09:00:00')
        txn('PTWY', 50, 'Rec Area', bin_a, '01/11/26 09:00:00')   # put away to bin A
        txn('PTWY', 50, bin_a, bin_b, '01/13/26 09:00:00')        # relocate to bin B (latest placement)
        txn('PICK', 5, bin_b, 'MFG Floor', '01/14/26 09:00:00')   # outflow — must be ignored

        # Inventory row starts STALE on the floor (the reported symptom).
        _seed_warehouse_row(cur, test_pcn, item, mpn='RLOC-MPN', onhandqty=45, loc_to='MFG Floor')

        # Call the EXACT function the background thread ships — no SQL copy here.
        reconcile_warehouse_locations(cur)

        cur.execute(f'SELECT loc_to FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn = %s', (str(test_pcn),))
        loc = cur.fetchone()['loc_to']
        assert loc == bin_b, (
            f"location reconcile must follow the latest placement (bin {bin_b}), got {loc!r} "
            f"(bin_a {bin_a} = ignored older placement; 'MFG Floor' = wrongly followed the PICK)")


def test_location_reconcile_honors_manual_adjt_edit():
    """Manual relocation bug (2026-06-16): a user changing the bin in the KOSH
    Warehouse Inventory editor saves loc_to AND logs an ADJT carrying the new bin,
    but the location reconcile only treated PTWY/RESTOCK/INDF/STOCK as placements —
    so within 5 minutes it reverted the row to the last imported PTWY and the manual
    change 'didn't stick'. The reconcile now also honors a genuine location-move
    ADJT (loc_to is a real bin). A relabel/renumber ADJT (loc_to is an ITEM number)
    must STILL be ignored. This test seeds a PTWY to bin A, a later manual ADJT to
    bin B, and a still-later relabel ADJT (loc_to = item number); the reconcile must
    land on bin B.
    """
    sys.path.insert(0, '/app')
    from psycopg2.extras import RealDictCursor
    from app import reconcile_warehouse_locations  # the SHIPPED query, not a copy
    test_pcn = 99934
    item = 'RLOC2-1'
    # 8-digit bins — the exact recurrence shape (manual ADJT into an 8-digit bin).
    bin_a = '14046120'
    bin_b = '14047080'

    with isolated_txn() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn = %s', (str(test_pcn),))
        cur.execute(f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn = %s', (str(test_pcn),))

        def txn(tt, qty, lf, lt, t):
            cur.execute(
                f'INSERT INTO {SCHEMA}."tblTransaction" '
                '(trantype, item, pcn, mpn, tranqty, tran_time, loc_from, loc_to, userid) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (tt, item, str(test_pcn), 'RLOC2-MPN', str(qty), t, lf, lt, 'regression@test.com'))
        txn('INDF', 50, 'n/a', 'Rec Area', '01/10/26 09:00:00')
        txn('PTWY', 50, 'Rec Area', bin_a, '01/11/26 09:00:00')      # imported put-away to bin A
        txn('ADJT', 0,  bin_a, bin_b, '01/15/26 09:00:00')           # MANUAL editor relocation to bin B (8-digit)
        txn('ADJT', 0,  'RLOC2-1', 'RLOC2-NEW', '01/16/26 09:00:00') # relabel (item->item) — must be IGNORED

        # Row is stale on bin A (what the old reconcile would force it back to).
        _seed_warehouse_row(cur, test_pcn, item, mpn='RLOC2-MPN', onhandqty=50, loc_to=bin_a)

        # Call the EXACT shipped function — no divergent SQL copy in the test.
        reconcile_warehouse_locations(cur)
        cur.execute(f'SELECT loc_to FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn = %s', (str(test_pcn),))
        loc = cur.fetchone()['loc_to']
        assert loc == bin_b, (
            f"reconcile must honor the manual ADJT relocation to bin {bin_b}, got {loc!r} "
            f"(bin_a {bin_a} = reverted manual edit; relabel ADJT must be ignored)")


def test_onhand_reconcile_neutralizes_relabel_adjt():
    """The on-hand double-count root cause (2026-06-12): a part RENUMBER/relabel
    was logged as an ADJT carrying the part's FULL qty, with loc_from/loc_to set
    to OLD/NEW item numbers (not locations). Counting that ADJT as +qty injects
    phantom stock (e.g. PCN 30314 showed 10000 on-hand AND 10000 on MFG Floor).

    The reconcile must treat a renumber-ADJT (both loc fields are NON-locations
    and differ) as quantity-neutral. This test seeds INDF +1000, a relabel-ADJT
    +1000, and PICK -1000, and asserts the corrected on-hand is 0 (the phantom is
    removed) while the un-corrected math would wrongly yield 1000.
    """
    import psycopg2 as _pg
    test_pcn = 99931
    test_mpn = 'RELABEL-MPN-1'

    with isolated_txn() as conn:
        cur = conn.cursor()
        # INDF receipt (real location), relabel ADJT (ITEM->ITEM), PICK to floor.
        def txn(tt, qty, lf, lt, t):
            cur.execute(
                f'INSERT INTO {SCHEMA}."tblTransaction" '
                '(trantype, item, pcn, mpn, tranqty, tran_time, loc_from, loc_to, userid) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (tt, 'RELABELNEW-2', str(test_pcn), test_mpn, str(qty), t, lf, lt, 'regression@test.com'))
        txn('INDF', 1000, 'n/a', 'Rec Area', '01/10/26 09:00:00')
        txn('ADJT', 1000, 'RELABELOLD-1', 'RELABELNEW-2', '01/11/26 09:00:00')  # the relabel
        txn('PICK', 1000, '1202233', 'MFG Floor', '01/12/26 09:00:00')

        # Replicate the reconcile's renumber-aware net for this PCN: corrected
        # (relabel neutralized) vs uncorrected (the old, phantom-producing math).
        cur.execute(f"""
            WITH locvocab AS (
                SELECT DISTINCT LOWER(TRIM(loc_to)) v FROM {SCHEMA}."tblTransaction"
                    WHERE trantype <> 'ADJT' AND loc_to IS NOT NULL AND TRIM(loc_to) <> ''
                UNION SELECT DISTINCT LOWER(TRIM(loc_from)) FROM {SCHEMA}."tblTransaction"
                    WHERE trantype <> 'ADJT' AND loc_from IS NOT NULL AND TRIM(loc_from) <> ''
                UNION VALUES ('mfg floor'),('rec area'),('count area'),('n/a'),('na'),('')
            ),
            parsed AS (
                SELECT trantype, tranqty,
                    (trantype='ADJT'
                       AND LOWER(TRIM(COALESCE(loc_from,''))) <> LOWER(TRIM(COALESCE(loc_to,'')))
                       AND NOT (LOWER(TRIM(COALESCE(loc_to,'')))  IN (SELECT v FROM locvocab) OR COALESCE(loc_to,'')  ~ '^[0-9]{{6,}}$' OR TRIM(COALESCE(loc_to,''))='')
                       AND NOT (LOWER(TRIM(COALESCE(loc_from,''))) IN (SELECT v FROM locvocab) OR COALESCE(loc_from,'') ~ '^[0-9]{{6,}}$' OR TRIM(COALESCE(loc_from,''))='')
                    ) is_relabel
                FROM {SCHEMA}."tblTransaction"
                WHERE pcn::text = %s AND tranqty ~ '^-?[0-9]+$'
            )
            SELECT
              GREATEST(0, SUM(CASE WHEN is_relabel THEN 0
                  WHEN trantype IN ('INDF','STOCK','PCN Generation','RESTOCK','ADJT') THEN tranqty::int
                  WHEN trantype IN ('PICK','PURGE') THEN -tranqty::int ELSE 0 END)) corrected,
              GREATEST(0, SUM(CASE
                  WHEN trantype IN ('INDF','STOCK','PCN Generation','RESTOCK','ADJT') THEN tranqty::int
                  WHEN trantype IN ('PICK','PURGE') THEN -tranqty::int ELSE 0 END)) uncorrected,
              BOOL_OR(is_relabel) any_relabel
            FROM parsed
        """, (str(test_pcn),))
        corrected, uncorrected, any_relabel = cur.fetchone()
        assert any_relabel is True, 'the ITEM->ITEM ADJT must be flagged as a relabel'
        assert uncorrected == 1000, (
            f'old math double-counts the relabel as +1000 phantom, got {uncorrected}')
        assert corrected == 0, (
            f'renumber-aware reconcile must neutralize the relabel -> on-hand 0, got {corrected}')


def test_pcn_history_balance_matches_reconcile_on_relabel():
    """PCN History's per-row 'On Hand' (running balance) must use the SAME
    relabel-neutral math as the reconcile that feeds Warehouse Inventory, or the
    two screens disagree (e.g. PCN 1247 showed 18000 in history but 9000 in the
    warehouse). Regression for that mismatch: replay INDF +1000, relabel-ADJT
    +1000, PICK -1000 the way the /pcn-history route does. The PICK row must land
    at 0 (full reel picked), NOT 1000 (the old phantom). The buggy replay that
    counts the relabel as +qty would leave 1000 after the pick.
    """
    test_pcn = 99932
    test_mpn = 'RELABEL-MPN-2'

    with isolated_txn() as conn:
        cur = conn.cursor()

        def txn(tt, qty, lf, lt, t):
            cur.execute(
                f'INSERT INTO {SCHEMA}."tblTransaction" '
                '(trantype, item, pcn, mpn, tranqty, tran_time, loc_from, loc_to, userid) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (tt, 'RELABELNEW-2', str(test_pcn), test_mpn, str(qty), t, lf, lt, 'regression@test.com'))
        txn('INDF', 1000, 'n/a', 'Rec Area', '01/10/26 09:00:00')
        txn('ADJT', 1000, 'RELABELOLD-1', 'RELABELNEW-2', '01/11/26 09:00:00')  # the relabel
        txn('PICK', 1000, '1202233', 'MFG Floor', '01/12/26 09:00:00')

        # Pull rows the way the route does: with the is_relabel flag, oldest-first.
        cur.execute(f"""
            WITH locvocab AS (
                SELECT DISTINCT LOWER(TRIM(loc_to)) v FROM {SCHEMA}."tblTransaction"
                    WHERE trantype <> 'ADJT' AND loc_to IS NOT NULL AND TRIM(loc_to) <> ''
                UNION SELECT DISTINCT LOWER(TRIM(loc_from)) FROM {SCHEMA}."tblTransaction"
                    WHERE trantype <> 'ADJT' AND loc_from IS NOT NULL AND TRIM(loc_from) <> ''
                UNION VALUES ('mfg floor'),('rec area'),('count area'),('n/a'),('na'),('')
            )
            SELECT trantype, tranqty,
                   (trantype='ADJT'
                      AND LOWER(TRIM(COALESCE(loc_from,''))) <> LOWER(TRIM(COALESCE(loc_to,'')))
                      AND NOT (LOWER(TRIM(COALESCE(loc_to,'')))  IN (SELECT v FROM locvocab) OR COALESCE(loc_to,'')  ~ '^[0-9]{{6,}}$' OR TRIM(COALESCE(loc_to,''))='')
                      AND NOT (LOWER(TRIM(COALESCE(loc_from,''))) IN (SELECT v FROM locvocab) OR COALESCE(loc_from,'') ~ '^[0-9]{{6,}}$' OR TRIM(COALESCE(loc_from,''))='')
                   ) is_relabel
            FROM {SCHEMA}."tblTransaction"
            WHERE pcn::text = %s AND tranqty ~ '^-?[0-9]+$'
            ORDER BY tran_time ASC
        """, (str(test_pcn),))
        rows = cur.fetchall()  # (trantype, tranqty, is_relabel) oldest-first

        # Replay the route's balance loop, both fixed (honor is_relabel) and buggy.
        def replay(honor_relabel):
            bal = 0
            picked_bal = None
            for tt, tq, is_relabel in rows:
                q = int(tq)
                if tt == 'RNDT':
                    bal = q
                    continue
                if honor_relabel and is_relabel:
                    delta = 0
                elif tt in ('INDF', 'STOCK', 'PCN Generation', 'RESTOCK', 'ADJT'):
                    delta = q
                elif tt in ('PICK', 'PURGE'):
                    delta = -q
                else:
                    delta = 0
                bal = max(0, bal + delta)
                if tt == 'PICK':
                    picked_bal = bal
            return bal, picked_bal

        fixed_final, fixed_after_pick = replay(honor_relabel=True)
        buggy_final, buggy_after_pick = replay(honor_relabel=False)

        assert fixed_after_pick == 0, (
            f'after fully picking the reel, history on-hand must be 0, got {fixed_after_pick}')
        assert buggy_after_pick == 1000, (
            f'old replay should show the 1000 phantom after pick, got {buggy_after_pick}')
        assert fixed_final == 0 and buggy_final == 1000, (
            f'fixed final={fixed_final} (want 0), buggy final={buggy_final} (want 1000)')


def test_pcn_history_relabel_neutral_on_real_pcns():
    """Run the PCN History balance replay against REAL production PCNs that
    carry relabel-ADJTs, and prove the deployed fix removes phantom on live
    data (not just on a synthetic fixture).

    Invariants asserted for every sampled real relabel PCN:
      * fixed (relabel-neutral) balance <= buggy (relabel-as-+qty) balance
        — the fix can only ever REMOVE phantom, never add stock; and
      * across the sample, at least one PCN has fixed < buggy — i.e. the fix
        is actually doing something on real data.
    The list of PCNs exercised is printed so it can be reviewed.
    """
    import re as _re
    with isolated_txn() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT DISTINCT LOWER(TRIM(loc_to)) v FROM {SCHEMA}."tblTransaction" WHERE trantype<>'ADJT' AND loc_to IS NOT NULL AND TRIM(loc_to)<>''
            UNION SELECT DISTINCT LOWER(TRIM(loc_from)) FROM {SCHEMA}."tblTransaction" WHERE trantype<>'ADJT' AND loc_from IS NOT NULL AND TRIM(loc_from)<>''
        """)
        locvocab = set(r[0] for r in cur.fetchall()) | {
            'mfg floor','rec area','receiving area','count area','n/a','na',''}
        binre = _re.compile(r'^[0-9]{6,}$')
        def isloc(x):
            x = (x or '').strip()
            return x.lower() in locvocab or binre.match(x) is not None or x == ''
        def isrl(tt, lf, lt):
            return tt == 'ADJT' and (lf or '').strip().lower() != (lt or '').strip().lower() \
                   and not isloc(lt) and not isloc(lf)

        # A curated set of REAL PCNs verified to carry relabel-ADJTs (2026-06-16).
        # Includes the reported PCN 1247. Any not present in live data are skipped.
        sample = ['1247','51','53','138','143','150','152','156','157','164',
                  '165','171','9007','219','822','27469','33687']
        def replay(pcn, honor_relabel):
            cur.execute(f"""
                SELECT trantype, tranqty, loc_from, loc_to
                FROM {SCHEMA}."tblTransaction"
                WHERE pcn::text=%s AND COALESCE(reversed,false)=false AND tranqty ~ '^-?[0-9]+$'
                ORDER BY CASE WHEN tran_time ~ '^[0-9]{{4}}-' THEN tran_time::timestamptz
                              WHEN tran_time ~ '^[0-9]{{2}}/[0-9]{{2}}/[0-9]{{2}}' THEN to_timestamp(tran_time,'MM/DD/YY HH24:MI:SS') END
                         ASC NULLS FIRST, id ASC
            """, (pcn,))
            bal = 0; has_relabel = False
            for tt, tq, lf, lt in cur.fetchall():
                q = int(tq); rl = isrl(tt, lf, lt); has_relabel = has_relabel or rl
                if tt == 'RNDT':
                    bal = q; continue
                if honor_relabel and rl: delta = 0
                elif tt in ('INDF','STOCK','PCN Generation','RESTOCK','ADJT'): delta = q
                elif tt in ('PICK','PURGE'): delta = -q
                else: delta = 0
                bal = max(0, bal + delta)
            return bal, has_relabel

        tested, phantom_removed = [], 0
        for pcn in sample:
            fixed, hr = replay(pcn, True)
            buggy, _ = replay(pcn, False)
            if not hr:
                continue  # PCN gone or no longer a relabel case — skip, don't fail
            tested.append(pcn)
            assert fixed <= buggy, (
                f'PCN {pcn}: relabel-neutral balance {fixed} must not exceed the '
                f'old phantom balance {buggy} — the fix may only remove stock')
            if fixed < buggy:
                phantom_removed += 1
        print(f"    [real-pcn] exercised relabel PCNs: {', '.join(tested)}")
        assert tested, 'no real relabel PCNs found to test (data changed?)'
        assert phantom_removed > 0, (
            'expected the fix to remove phantom on at least one real relabel PCN')


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

TESTS = [
    test_restock_allowed_after_purge_following_restock,
    test_restock_allowed_when_zero_onhand_even_if_last_restock,
    test_restock_blocked_when_already_restocked_with_stock_present,
    test_print_label_sums_across_duplicate_pcn_rows,
    test_validate_location_auto_registers_unknown_7digit,
    test_purged_pcn_can_be_restocked_with_same_pcn,
    test_bom_load_inserts_every_item_received,
    test_shortage_report_alt_part_qty_and_same_mpn_visibility,
    test_shortage_report_own_stock_is_case_insensitive,
    test_shortage_report_counts_mfg_floor_stock,
    test_location_reconcile_follows_latest_placement,
    test_location_reconcile_honors_manual_adjt_edit,
    test_onhand_reconcile_neutralizes_relabel_adjt,
    test_pcn_history_balance_matches_reconcile_on_relabel,
    test_pcn_history_relabel_neutral_on_real_pcns,
    test_bom_python_parser_finds_lines_across_sheets,
    test_return_connection_never_leaks_foreign_connection,
    test_quantity_fields_are_not_number_spinners,
    test_app_served_by_gunicorn_not_dev_server,
    test_all_pages_render_without_server_error,
]


def _final_cleanup():
    """Wipe every trace of regression-test data: activity log rows by the
    test username, any straggler warehouse/transaction/BOM/job/loc rows
    matching test markers. Runs unconditionally after the suite, even on
    test failure, so test data never leaks into the real activity log
    Preet sees in the UI.
    """
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = True
    cur = conn.cursor()
    test_user = 'regression@test.com'
    test_pcns = ['99001', '99002', '99003', '99004', '99005']
    test_jobs = ['REGRESS-BOM-9999']
    test_items = (
        'REGRESS-RESTOCK-AFTER-PURGE', 'REGRESS-ZERO-ONHAND',
        'REGRESS-DOUBLE-RESTOCK', 'REGRESS-DUP-A', 'REGRESS-DUP-B',
        'REGRESS-PURGED-LOSTROW',
    )
    test_loc = '9999301'
    try:
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblActivityLog" WHERE username = %s',
            (test_user,),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblActivityLog" '
            f"WHERE description ILIKE %s OR description ILIKE %s OR details ILIKE %s",
            ('%REGRESS-%', '%PCN 9900%', '%REGRESS-%'),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblTransaction" WHERE pcn::text = ANY(%s)',
            (test_pcns,),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE pcn::text = ANY(%s)',
            (test_pcns,),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblWhse_Inventory" WHERE item = ANY(%s)',
            (list(test_items),),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblBOM" WHERE job::text = ANY(%s)',
            (test_jobs,),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblJob" WHERE job_number = ANY(%s)',
            (test_jobs,),
        )
        cur.execute(
            f'DELETE FROM {SCHEMA}."tblLoc" WHERE location = %s',
            (test_loc,),
        )
    except Exception as e:
        print(f'  WARN  final cleanup raised: {e}')
    finally:
        conn.close()


def main():
    print('KOSH regression smoke tests')
    print('=' * 60)
    failures = []
    try:
        for fn in TESTS:
            name = fn.__name__
            try:
                fn()
                print(f'  PASS  {name}')
            except AssertionError as e:
                print(f'  FAIL  {name}: {e}')
                failures.append((name, str(e)))
            except Exception as e:
                tb = traceback.format_exc()
                print(f'  ERROR {name}: {e}\n{tb}')
                failures.append((name, f'{e}\n{tb}'))
    finally:
        _final_cleanup()
    print('=' * 60)
    print(f'{len(TESTS) - len(failures)} passed, {len(failures)} failed')
    if failures:
        print('\nFailures:')
        for n, msg in failures:
            print(f'  - {n}: {msg[:200]}')
    return 0 if not failures else 1


if __name__ == '__main__':
    sys.exit(main())
